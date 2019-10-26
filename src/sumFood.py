import openpyxl
import time,sys, re, os
from collections import defaultdict

current_dir = os.getcwd()
result = defaultdict(int)
regular_foods = input('Do you want to sum foods (yes/no)?\n')
os.system('cls')
extra_foods = input('Do you want to sum extra foods (yes/no)?\n')

# Sum book in current directory
for i in os.listdir():
        if '.xls' in i and 'ΤΡΟΦΙΜΑ' in i:
                bookname = i
# Loading excel book
book = openpyxl.load_workbook(bookname)

#Getting only needed sheets
sheets = book.sheetnames
paketa = []
sheet_names = {}
working_sheets = []

# Loading sheets
for i in sheets:
        if 'ΠΑΚΕΤΟ' in i:
                paketa.append(i)

# Creating sheetnames keys (sheet1, sheet2, etc.)
for index, sheet in enumerate(paketa):
        sheet_names['sheet'+str(index)] = sheet

# Appending working sheets to array
for index, sheet in enumerate(paketa):
        working_sheets.append(book[paketa[index]])

 # Sum ONE sheet     
def sumSheet(worksheet):
        global result

        rows = worksheet.max_row
        table = []
        foods = []

        if regular_foods == 'yes':
                # Creating sheet table
                for i in range(3, rows):
                        val = worksheet.cell(row=i, column=4).value
                        # Creating rows: Appending cells with their values to table array if not empty.
                        if val:
                                val = re.sub(r"(\d),(\d)", r"\1.\2", val)
                                table.append(val.strip(', ').split(', '))
        if extra_foods == 'yes':
                # Creating sheet table
                for i in range(3, rows):
                        val = worksheet.cell(row=i, column=5).value
                        # Creating rows: Appending cells with their values to table array if not empty.
                        if val:
                                val = re.sub(r"(\d),(\d)", r"\1.\2", val)
                                table.append(val.strip(', ').split(', '))
                        

        # Loop through table rows
        for row in (table):
                # Loop through each row item             
                for item in row:
                        # Get quantity
                        quantity = item.strip()[0]
                        # Get food after stripping it
                        food = item.strip()[2:].strip() if item.strip()[0].isdigit() else item.strip()[0:]
                        
                        #If there is quantity
                        if type(item[0]) == int:
                                #Append food and quantity dictionary
                                foods.append({'food':food,'quantity':quantity})
                        #If absent
                        elif item == 'ΔΕΝ ΗΡΘΕ':
                                foods.append({'food':'ΔΕΝ ΗΡΘΕ','quantity':1})
                        #Else (i.e. 'ΑΛΕΥΡΙ')
                        else:
                                #Append food and quantity dictionary
                                foods.append({'food':food,'quantity':1})

        # Sum values for each food
        for record in foods:
                result[record['food']] += int(record['quantity'])

        
def sumBook():        
        # Performing calculation for every sheet in workbook
        for index, sheet in enumerate(working_sheets):
                sumSheet(sheet)
                #Convert food-quantity pair to a dict within a list
                final_result = [{'food': food, 'quantity': quantity} for food, quantity in result.items()]
        for index, d in enumerate(result):

                print(final_result[index]['food'] + ': ' + str(final_result[index]['quantity']) +'\n')


sumBook()
k = input('\nPress any key to exit.')